#!/usr/bin/env python3
"""Small JSON API helper for editing exam-data-master.xlsx."""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


APP_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = APP_ROOT / "exports" / "exam-data-master.xlsx"
SHEET_NAME = "tat_ca_cau"
BACKUP_DIR = APP_ROOT / "exports" / "backups"

VISIBLE_HEADERS = [
    "ma_de",
    "ten_de",
    "ky_nang",
    "part",
    "ten_part",
    "so_cau",
    "dang_cau_hoi",
    "loai_dap_an",
    "tieu_de",
    "noi_dung_bai_doc",
    "cau_hoi",
    "anh_neu_co",
    "audio",
    "lua_chon_A",
    "lua_chon_B",
    "lua_chon_C",
    "lua_chon_D",
    "lua_chon_E",
    "lua_chon_F",
    "lua_chon_G",
    "lua_chon_H",
    "dap_an_dung",
    "dap_an_chap_nhan",
    "giai_thich",
    "dong_goc",
    "ghi_chu",
]

EDITABLE_HEADERS = [
    "ten_part",
    "dang_cau_hoi",
    "loai_dap_an",
    "tieu_de",
    "noi_dung_bai_doc",
    "cau_hoi",
    "anh_neu_co",
    "audio",
    "lua_chon_A",
    "lua_chon_B",
    "lua_chon_C",
    "lua_chon_D",
    "lua_chon_E",
    "lua_chon_F",
    "lua_chon_G",
    "lua_chon_H",
    "dap_an_dung",
    "dap_an_chap_nhan",
    "giai_thich",
    "dong_goc",
    "ghi_chu",
]


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def load_sheet():
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Cannot find {WORKBOOK}")
    workbook = load_workbook(WORKBOOK)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Cannot find sheet {SHEET_NAME}")
    sheet = workbook[SHEET_NAME]
    headers = [cell_text(cell.value) for cell in sheet[1]]
    positions = {header: index + 1 for index, header in enumerate(headers) if header}
    return workbook, sheet, headers, positions


def row_to_dict(sheet, headers: list[str], row_number: int) -> dict[str, str]:
    row = {"rowId": str(row_number)}
    for index, header in enumerate(headers, start=1):
        if header in VISIBLE_HEADERS:
            row[header] = cell_text(sheet.cell(row=row_number, column=index).value)
    return row


def list_rows() -> dict[str, Any]:
    workbook, sheet, headers, _positions = load_sheet()
    rows = []
    for row_number in range(2, sheet.max_row + 1):
        row = row_to_dict(sheet, headers, row_number)
        if any(row.get(header, "").strip() for header in VISIBLE_HEADERS):
            rows.append(row)
    workbook.close()

    return {
        "ok": True,
        "workbook": str(WORKBOOK),
        "headers": [header for header in VISIBLE_HEADERS if header in headers],
        "editableHeaders": [header for header in EDITABLE_HEADERS if header in headers],
        "rows": rows,
        "filters": {
            "exams": sorted({row.get("ma_de", "") for row in rows if row.get("ma_de")}),
            "skills": sorted({row.get("ky_nang", "") for row in rows if row.get("ky_nang")}),
            "parts": sorted({row.get("part", "") for row in rows if row.get("part")}, key=part_sort_key),
        },
    }


def part_sort_key(value: str) -> tuple[int, str]:
    return (0, f"{int(value):03d}") if str(value).isdigit() else (1, str(value))


def make_backup() -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    target = BACKUP_DIR / f"exam-data-master-{stamp}.xlsx"
    shutil.copy2(WORKBOOK, target)
    return target


def update_row(payload: dict[str, Any]) -> dict[str, Any]:
    row_id = int(str(payload.get("rowId") or "0"))
    values = payload.get("values") or {}
    if row_id < 2:
        raise ValueError("rowId không hợp lệ.")

    workbook, sheet, headers, positions = load_sheet()
    if row_id > sheet.max_row:
        raise ValueError("Không tìm thấy dòng cần sửa trong Excel.")

    changed: dict[str, dict[str, str]] = {}
    for header, value in values.items():
        if header not in EDITABLE_HEADERS or header not in positions:
            continue
        column = positions[header]
        cell = sheet.cell(row=row_id, column=column)
        before = cell_text(cell.value)
        after = cell_text(value)
        if before != after:
            cell.value = after
            changed[header] = {"before": before, "after": after}

    if not changed:
        workbook.close()
        return {"ok": True, "changed": {}, "message": "Không có thay đổi."}

    backup = make_backup()
    sync_matching_exam_sheet(workbook, sheet, headers, positions, row_id, changed)
    try:
        workbook.save(WORKBOOK)
    except PermissionError as exc:
        workbook.close()
        raise PermissionError("Excel đang mở file này. Hãy lưu và đóng Excel rồi thử lại.") from exc
    workbook.close()

    return {
        "ok": True,
        "rowId": row_id,
        "changed": changed,
        "backup": str(backup),
        "row": read_single_row(row_id),
    }


def sync_matching_exam_sheet(workbook, master_sheet, headers, positions, row_id: int, changed: dict[str, Any]) -> None:
    exam_id = cell_text(master_sheet.cell(row=row_id, column=positions.get("ma_de", 0)).value)
    skill = cell_text(master_sheet.cell(row=row_id, column=positions.get("ky_nang", 0)).value)
    part = cell_text(master_sheet.cell(row=row_id, column=positions.get("part", 0)).value)
    question = cell_text(master_sheet.cell(row=row_id, column=positions.get("so_cau", 0)).value)
    if not exam_id or exam_id not in workbook.sheetnames:
        return

    sheet = workbook[exam_id]
    sheet_headers = [cell_text(cell.value) for cell in sheet[1]]
    sheet_positions = {header: index + 1 for index, header in enumerate(sheet_headers) if header}
    required = ("ma_de", "ky_nang", "part", "so_cau")
    if any(header not in sheet_positions for header in required):
        return

    for candidate_row in range(2, sheet.max_row + 1):
        if (
            cell_text(sheet.cell(row=candidate_row, column=sheet_positions["ma_de"]).value) == exam_id
            and cell_text(sheet.cell(row=candidate_row, column=sheet_positions["ky_nang"]).value) == skill
            and cell_text(sheet.cell(row=candidate_row, column=sheet_positions["part"]).value) == part
            and cell_text(sheet.cell(row=candidate_row, column=sheet_positions["so_cau"]).value) == question
        ):
            for header, diff in changed.items():
                if header in sheet_positions:
                    sheet.cell(row=candidate_row, column=sheet_positions[header]).value = diff["after"]
            return


def read_single_row(row_id: int) -> dict[str, str]:
    workbook, sheet, headers, _positions = load_sheet()
    row = row_to_dict(sheet, headers, row_id)
    workbook.close()
    return row


def read_stdin_json() -> dict[str, Any]:
    raw = sys.stdin.read().strip() or "{}"
    return json.loads(raw)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("action", choices=["list", "update"])
    args = parser.parse_args()

    try:
        if args.action == "list":
            payload = list_rows()
        else:
            payload = update_row(read_stdin_json())
    except Exception as exc:
        payload = {"ok": False, "error": str(exc)}
        print(json.dumps(payload, ensure_ascii=False))
        sys.exit(1)

    print(json.dumps(payload, ensure_ascii=False))


if __name__ == "__main__":
    main()
